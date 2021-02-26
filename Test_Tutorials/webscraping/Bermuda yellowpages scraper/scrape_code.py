from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup as bs


def get_ids(content):
    # ### BUSINESS NAMES ###
    soup = bs(content.content, features='html.parser')
    title = soup.find(attrs={'class': 'block-title'}).h3.string

    ### ADDRESS ####
    address = soup.find_all('span', attrs={'itemprop': [
        'streetAddress',
        'postOfficeBoxNumber',
        'addressRegion',
        'addressCountry',
        'postalCode'

    ]})
    address = [text.get_text() for text in address]

    # ### PHONE NUMBER ###
    phone_number = soup.find_all(attrs={'class': 'box-ico phone-i'})

    phone_number = [call.a.get('href')[4:] for call in phone_number]

    # ### DESCRIPTION ###
    description = soup.find(attrs={'class': 'description'})
    description = description.find_all('p')
    description = [d.get_text() for d in description]

    # ### TAGS ###
    tags = soup.find_all(attrs={'class': ("main-tags", 'secondary-tags')})
    tags = [b.get_text() for i in tags for b in i.find_all('a')]

    # ### LOGO ###
    try:
        logo = soup.find(attrs={'class': 'logo'}).get('src')
    except AttributeError:
        logo = ''


    business = {'Name'         : title,
                'Address'      : address,
                'Phone Numbers': phone_number,
                'Description'  : description,
                'Tags'         : tags,
                'Logo'         : logo
                }
    return business


def save_excel_book(workbook, workbook_name='file', dest=''):
    xlbook = Path(f"{dest}{workbook_name}.xlsx")

    workbook.save(xlbook)


def store_to_database(list_of_businesses, wb=None, category=None):
    if wb is None:
        wb = openpyxl.Workbook()
    if category is None:
        ws = wb.active
    else:
        try:
            ws = wb[category]
        except KeyError:
            ws = wb.create_sheet(category)
    headers = ['Name', 'Address', 'Phone Numbers', 'Description', 'Tags', 'Logo']
    if not len(ws[1]) == len(headers):
        ws.append(headers)
    for line in list_of_businesses:
        # create generator to yield product value
        # use headers in desired order as `key`
        values = (line[k] for k in headers)
        l = []
        for v in values:
            if isinstance(v, list):
                collected = ', '.join(v)
                l.append(collected)
            else:
                l.append(v)
        ws.append(l)


def open_workbook(workbook_name='file.xlsx', dest='', sheet_title='sheet', header=None):
    if header is None:
        header = []
    xlbook = Path(f"{dest}{workbook_name}.xlsx")
    if xlbook.exists():
        workbook = openpyxl.load_workbook(xlbook.as_posix())
        return workbook
    else:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = sheet_title
        ws.append(header)
        return workbook